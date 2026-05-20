import csv
import os
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable, TypedDict

from dotenv import load_dotenv

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pptx import Presentation
from PyPDF2 import PdfReader

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_qdrant import QdrantVectorStore
from langchain_google_genai import ChatGoogleGenerativeAI

from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams

from langgraph.graph import StateGraph, START, END


BASE_DIR = Path(__file__).resolve().parent

load_dotenv(BASE_DIR / ".env")

DRIVE_FOLDER_ID = "1bljckgh6CFCPpB7dHfoeMBdNx3Il7gWX"

TOKEN_PATH = BASE_DIR / "google_token.json"
CREDENTIALS_PATH = (
    BASE_DIR
    / "client_secret_747615723880-g0e74bm27ifn75h231po8tb3a6mqaodt.apps.googleusercontent.com.json"
)

GOOGLE_DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

EMBEDDING_MODEL = "sentence-transformers/all-MiniLM-L6-v2"
EMBEDDING_VECTOR_SIZE = 384

QDRANT_URL = os.getenv("QDRANT_URL")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY")
QDRANT_COLLECTION_NAME = os.getenv("QDRANT_COLLECTION_NAME", "predikly_t7")

FOLDER_MIME = "application/vnd.google-apps.folder"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PPTX_MIME = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
REFERENCE_EXCEL_NAME = "PREDIKLY_USECASE_REFRENCE_POINTS.xlsx"
REFERENCE_EXCEL_BASENAME = "PREDIKLY_USECASE_REFRENCE_POINTS"
LOCAL_REFERENCE_EXCEL_PATH = BASE_DIR / REFERENCE_EXCEL_NAME


def normalize_match_key(value: str) -> str:
    value = value.lower().strip()
    value = value.replace(".pptx", "")
    value = value.replace(".xlsx", "")
    value = value.replace("’", "'")
    value = value.replace("–", "-")
    value = value.replace("—", "-")
    value = re.sub(r"\s+", " ", value)
    return value


def map_excel_rows_to_drive_ppts_exact(
    excel_rows: list[dict],
    drive_files: list[dict],
    *,
    excel_match_column: str,
) -> tuple[list[dict], list[dict]]:
    ppt_files = [
        file
        for file in drive_files
        if file.get("mimeType") == PPTX_MIME
    ]

    ppt_by_key = {}

    for ppt in ppt_files:
        key = normalize_match_key(ppt["name"])

        if key in ppt_by_key:
            raise ValueError(
                f"Duplicate PPT match key found: {key}. "
                f"Files: {ppt_by_key[key]['name']} and {ppt['name']}"
            )

        ppt_by_key[key] = ppt

    mapped_items = []
    unmatched_rows = []

    for row in excel_rows:
        raw_value = row.get(excel_match_column)

        if not raw_value:
            unmatched_rows.append(
                {
                    "reason": "missing_excel_match_value",
                    "excel_row": row,
                }
            )
            continue

        row_key = normalize_match_key(str(raw_value))
        ppt = ppt_by_key.get(row_key)

        if not ppt:
            unmatched_rows.append(
                {
                    "reason": "no_exact_ppt_match",
                    "match_key": row_key,
                    "excel_row": row,
                }
            )
            continue

        mapped_items.append(
            {
                "excel_row": row,
                "drive_file": ppt,
                "match_key": row_key,
            }
        )

    return mapped_items, unmatched_rows


def find_unmapped_drive_ppts(
    mapped_items: list[dict],
    drive_files: list[dict],
) -> list[dict]:
    mapped_drive_ids = {
        item["drive_file"]["id"]
        for item in mapped_items
    }

    return [
        file
        for file in drive_files
        if file.get("mimeType") == PPTX_MIME
        and file["id"] not in mapped_drive_ids
    ]


def map_drive_ppts_to_excel_rows_exact(
    drive_files: list[dict],
    excel_rows: list[dict],
    *,
    excel_match_column: str,
) -> tuple[list[dict], list[dict], list[dict]]:
    excel_by_key = {}

    for row in excel_rows:
        raw_value = row.get(excel_match_column)

        if not raw_value:
            continue

        key = normalize_match_key(str(raw_value))
        excel_by_key.setdefault(key, []).append(row)

    mapped_items = []
    unmatched_ppts = []
    ppt_keys = set()

    for ppt in drive_files:
        if ppt.get("mimeType") != PPTX_MIME:
            continue

        ppt_key = normalize_match_key(ppt["name"])
        ppt_keys.add(ppt_key)
        matched_excel_rows = excel_by_key.get(ppt_key, [])

        if not matched_excel_rows:
            unmatched_ppts.append(
                {
                    "reason": "no_exact_excel_match",
                    "match_key": ppt_key,
                    "drive_file": ppt,
                }
            )
            continue

        mapped_items.append(
            {
                "drive_file": ppt,
                "excel_rows": matched_excel_rows,
                "match_key": ppt_key,
            }
        )

    extra_excel_rows = [
        {
            "reason": "excel_row_not_used_for_current_ppt_folder",
            "match_key": normalize_match_key(str(row.get(excel_match_column))),
            "excel_row": row,
        }
        for row in excel_rows
        if row.get(excel_match_column)
        and normalize_match_key(str(row.get(excel_match_column))) not in ppt_keys
    ]

    return mapped_items, unmatched_ppts, extra_excel_rows


def download_drive_file(service, file_id: str) -> bytes:
    request = service.files().get_media(fileId=file_id)
    buffer = BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False

    while not done:
        _, done = downloader.next_chunk()

    return buffer.getvalue()


def export_drive_file(service, file_id: str, mime_type: str) -> bytes:
    request = service.files().export_media(fileId=file_id, mimeType=mime_type)
    buffer = BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False

    while not done:
        _, done = downloader.next_chunk()

    return buffer.getvalue()


def extract_xlsx_rows(content: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [
        str(header).strip() if header is not None else f"column_{index}"
        for index, header in enumerate(rows[0], start=1)
    ]
    records = []

    for row in rows[1:]:
        record = {
            header: value
            for header, value in zip(headers, row)
        }

        if any(value is not None and str(value).strip() for value in record.values()):
            records.append(record)

    return records


def find_reference_excel_file(drive_files: list[dict]) -> dict:
    for file in drive_files:
        if file["name"] in {REFERENCE_EXCEL_NAME, REFERENCE_EXCEL_BASENAME}:
            return file

    raise ValueError(
        "Reference Excel file not found in the target Drive folder. "
        f"Expected: {REFERENCE_EXCEL_NAME} or {REFERENCE_EXCEL_BASENAME}"
    )


def search_reference_excel_file(service) -> dict:
    response = (
        service.files()
        .list(
            q=(
                f"name = '{REFERENCE_EXCEL_BASENAME}' "
                f"and trashed = false"
            ),
            pageSize=10,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            fields="files(id,name,mimeType,webViewLink,parents)",
        )
        .execute()
    )
    files = response.get("files", [])

    if not files:
        raise ValueError(
            "Reference Excel/Google Sheet was not found anywhere accessible. "
            f"Expected name: {REFERENCE_EXCEL_BASENAME}"
        )

    if len(files) > 1:
        names = ", ".join(f"{file['name']} ({file['id']})" for file in files)
        raise ValueError(f"Multiple reference sheets found: {names}")

    return files[0]


def load_reference_excel_rows(service, drive_files: list[dict]) -> list[dict]:
    if LOCAL_REFERENCE_EXCEL_PATH.exists():
        return extract_xlsx_rows(LOCAL_REFERENCE_EXCEL_PATH.read_bytes())

    try:
        reference_excel = find_reference_excel_file(drive_files)
    except ValueError:
        reference_excel = search_reference_excel_file(service)

    if reference_excel["mimeType"] == GOOGLE_SHEET_MIME:
        content = export_drive_file(service, reference_excel["id"], XLSX_MIME)
    elif reference_excel["mimeType"] == XLSX_MIME:
        content = download_drive_file(service, reference_excel["id"])
    else:
        raise ValueError(
            "Reference file must be a Google Sheet or XLSX file. "
            f"Got: {reference_excel['mimeType']}"
        )

    return extract_xlsx_rows(content)


def infer_excel_match_column(
    excel_rows: list[dict],
    drive_files: list[dict],
) -> str:
    if not excel_rows:
        raise ValueError("Reference Excel file has no data rows.")

    ppt_keys = {
        normalize_match_key(file["name"])
        for file in drive_files
        if file.get("mimeType") == PPTX_MIME
    }
    best_column = None
    best_score = 0

    for column in excel_rows[0].keys():
        score = 0

        for row in excel_rows:
            value = row.get(column)
            if value and normalize_match_key(str(value)) in ppt_keys:
                score += 1

        if score > best_score:
            best_score = score
            best_column = column

    if not best_column:
        columns = ", ".join(str(column) for column in excel_rows[0].keys())
        raise ValueError(
            "Could not infer the Excel column that exactly matches PPT names. "
            f"Available columns: {columns}"
        )

    return str(best_column)


def load_google_credentials() -> Credentials:
    creds = None

    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(
            str(TOKEN_PATH),
            GOOGLE_DRIVE_SCOPES,
        )

    if creds and creds.valid:
        return creds

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            str(CREDENTIALS_PATH),
            GOOGLE_DRIVE_SCOPES,
        )
        creds = flow.run_local_server(port=0)

    TOKEN_PATH.write_text(creds.to_json())
    return creds

def build_drive_service():
    credentials = load_google_credentials()
    return build("drive", "v3", credentials=credentials)

def get_drive_item(service, drive_id: str) -> dict:
    return (
        service.files()
        .get(
            fileId=drive_id,
            supportsAllDrives=True,
            fields="id,name,mimeType,webViewLink,parents",
        )
        .execute()
    )

def list_drive_folder(
    service,
    folder_id: str,
    *,
    recursive: bool = True,
    path: str = "",
) -> list[dict]:
    files = []
    page_token = None

    while True:
        response = (
            service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed = false",
                pageSize=1000,
                pageToken=page_token,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                fields="nextPageToken, files(id,name,mimeType,webViewLink,parents)",
            )
            .execute()
        )

        for item in response.get("files", []):
            item_path = f"{path}/{item['name']}" if path else item["name"]

            if item["mimeType"] == FOLDER_MIME:
                if recursive:
                    files.extend(
                        list_drive_folder(
                            service,
                            item["id"],
                            recursive=True,
                            path=item_path,
                        )
                    )
                continue

            item["drive_path"] = item_path
            files.append(item)

        page_token = response.get("nextPageToken")
        if not page_token:
            return files

def load_drive_file_metadata(service=None) -> list[dict]:
    service = service or build_drive_service()

    root_item = get_drive_item(service, DRIVE_FOLDER_ID)

    if root_item["mimeType"] == FOLDER_MIME:
        return list_drive_folder(
            service,
            root_item["id"],
            recursive=True,
            path=root_item["name"],
        )

    root_item["drive_path"] = root_item["name"]
    return [root_item]


def test_exact_excel_to_ppt_mapping() -> None:
    service = build_drive_service()
    drive_files = load_drive_file_metadata(service)
    excel_rows = load_reference_excel_rows(service, drive_files)
    excel_match_column = infer_excel_match_column(excel_rows, drive_files)

    mapped_items, unmatched_ppts, extra_excel_rows = map_drive_ppts_to_excel_rows_exact(
        drive_files=drive_files,
        excel_rows=excel_rows,
        excel_match_column=excel_match_column,
    )

    print(f"Excel match column: {excel_match_column}")
    print(f"Drive PPTs in target folder: {len([file for file in drive_files if file.get('mimeType') == PPTX_MIME])}")
    print(f"Excel rows in reference sheet: {len(excel_rows)}")
    print(f"Mapped PPTs: {len(mapped_items)}")
    print(f"Unmatched Drive PPTs: {len(unmatched_ppts)}")
    print(f"Extra Excel rows not used for this folder: {len(extra_excel_rows)}")

    for item in mapped_items:
        print(
            "OK:",
            item["drive_file"]["name"],
            "->",
            f"{len(item['excel_rows'])} Excel row(s)",
        )

    if unmatched_ppts:
        print("\nUnmatched Drive PPTs:")
        for item in unmatched_ppts:
            print(item["drive_file"]["drive_path"], "| match_key:", item["match_key"])

    if extra_excel_rows:
        print("\nExtra Excel rows not used for this folder:")
        for item in extra_excel_rows:
            print(item["excel_row"].get(excel_match_column))


if __name__ == "__main__":
    test_exact_excel_to_ppt_mapping()
